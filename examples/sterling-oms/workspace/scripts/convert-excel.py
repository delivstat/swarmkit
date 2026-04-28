"""Convert Excel files to markdown tables for RAG ingestion.

Excel rows without column context embed poorly. This script converts
each sheet to a markdown file with a proper table header, so the
vector search can match on column names + values together.

Prerequisites:
    pip install openpyxl

Usage:
    python scripts/convert-excel.py ~/sterling-knowledge/integrations/
    # Converts all .xlsx/.xls files in-place to .md alongside the originals
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)


def convert_excel_to_markdown(excel_path: Path, output_dir: Path) -> list[Path]:
    """Convert each sheet in an Excel file to a markdown table file."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    created: list[Path] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h or f"col_{i}") for i, h in enumerate(rows[0])]
        safe_name = f"{excel_path.stem}--{sheet_name}".replace(" ", "-").lower()
        out_file = output_dir / f"{safe_name}.md"

        lines = [
            f"# {excel_path.stem} — {sheet_name}",
            "",
            f"Source: {excel_path.name}",
            "",
            "| " + " | ".join(headers) + " |",
            "| " + " | ".join("---" for _ in headers) + " |",
        ]

        for row in rows[1:]:
            values = [str(v or "") for v in row]
            lines.append("| " + " | ".join(values) + " |")

        out_file.write_text("\n".join(lines), encoding="utf-8")
        created.append(out_file)

    wb.close()
    return created


def main(directory: Path) -> None:
    excel_files = list(directory.rglob("*.xlsx")) + list(directory.rglob("*.xls"))

    if not excel_files:
        print(f"No Excel files found in {directory}")
        return

    total = 0
    for ef in excel_files:
        created = convert_excel_to_markdown(ef, ef.parent)
        for f in created:
            print(f"  {f.name}")
        total += len(created)

    print(f"\nConverted {len(excel_files)} Excel files → {total} markdown files")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert Excel to markdown tables")
    parser.add_argument("directory", type=Path, help="Directory containing Excel files")
    args = parser.parse_args()
    main(args.directory)
