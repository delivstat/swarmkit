# /// script
# dependencies = [
#   "chromadb>=1.0",
#   "sentence-transformers[onnx]>=3.0",
#   "pypdf>=4.0",
#   "python-docx>=1.0",
#   "openpyxl>=3.1",
# ]
# ///
"""Ingest Sterling documentation into ChromaDB for RAG search.

Batch-embeds all documentation files using ONNX-accelerated
sentence-transformers. Ingests 20K files in 20-40 minutes on CPU
(vs 2+ days with one-at-a-time MCP ingestion).

Supported file types: .md, .txt, .pdf, .docx
HTML files are converted to text with link references preserved.

Prerequisites:
    uv (for PEP 723 inline dependency resolution)

Usage:
    export STERLING_DOCS_DIR=~/sterling-knowledge/product-docs
    uv run scripts/ingest-docs.py

    # Reset and re-ingest
    uv run scripts/ingest-docs.py --reset
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
import time
import warnings
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
logging.getLogger("pypdf").setLevel(logging.ERROR)

NATIVE_EXTENSIONS = {".md", ".txt"}
HTML_EXTENSIONS = {".html", ".htm"}
DOCX_EXTENSIONS = {".docx"}
PDF_EXTENSIONS = {".pdf"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
ALL_EXTENSIONS = (
    NATIVE_EXTENSIONS | HTML_EXTENSIONS | DOCX_EXTENSIONS | PDF_EXTENSIONS | EXCEL_EXTENSIONS
)
MAX_FILE_SIZE = 10_000_000  # 10MB
CHUNK_SIZE = 500
CHUNK_OVERLAP = 50
EMBED_BATCH_SIZE = 256


def _strip_html_preserve_links(html: str) -> str:
    """Convert HTML to text, preserving link references."""
    text = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL)
    text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)
    text = re.sub(
        r'<a\s+[^>]*href="([^"]*)"[^>]*>(.*?)</a>',
        r"\2 [ref: \1]",
        text,
        flags=re.DOTALL | re.IGNORECASE,
    )
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<p\s*/?>", "\n\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """Split text into overlapping chunks by words."""
    words = text.split()
    if len(words) <= chunk_size:
        return [text] if text.strip() else []
    chunks = []
    start = 0
    while start < len(words):
        end = start + chunk_size
        chunk = " ".join(words[start:end])
        if chunk.strip():
            chunks.append(chunk)
        start = end - overlap
    return chunks


def _read_pdf(path: Path) -> str:
    """Extract text from a PDF file."""
    from pypdf import PdfReader  # noqa: PLC0415

    reader = PdfReader(str(path))
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def _read_docx(path: Path) -> str:
    """Extract text from a DOCX file."""
    import docx  # noqa: PLC0415

    doc = docx.Document(str(path))
    return "\n\n".join(para.text for para in doc.paragraphs if para.text.strip())


def _read_excel(path: Path) -> str:
    """Extract text from an Excel file as markdown tables."""
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(h or f"col_{i}") for i, h in enumerate(rows[0])]
        lines = [
            f"## {sheet_name}",
            "",
            "| " + " | ".join(headers) + " |",
            "| " + " | ".join("---" for _ in headers) + " |",
        ]
        for row in rows[1:]:
            values = [str(v or "") for v in row]
            lines.append("| " + " | ".join(values) + " |")
        sheets.append("\n".join(lines))
    wb.close()
    return "\n\n".join(sheets)


def _read_file(path: Path, root: Path) -> str | None:
    """Read a file and return its text content."""
    suffix = path.suffix.lower()
    try:
        if suffix in PDF_EXTENSIONS:
            content = _read_pdf(path)
        elif suffix in DOCX_EXTENSIONS:
            content = _read_docx(path)
        elif suffix in EXCEL_EXTENSIONS:
            content = _read_excel(path)
        else:
            content = path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return None

    if suffix in HTML_EXTENSIONS:
        content = _strip_html_preserve_links(content)

    header = f"# {path.stem}\nSource: {path.relative_to(root)}\n\n"
    non_native = HTML_EXTENSIONS | PDF_EXTENSIONS | DOCX_EXTENSIONS | EXCEL_EXTENSIONS
    if suffix in non_native:
        content = header + content

    return content if content.strip() else None


def main() -> None:  # noqa: PLR0915
    parser = argparse.ArgumentParser(description="Batch ingest docs into ChromaDB")
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Delete existing collection and re-ingest from scratch",
    )
    args = parser.parse_args()

    base_dir = os.environ.get(
        "STERLING_DOCS_DIR",
        os.path.expanduser("~/sterling-knowledge"),
    )
    root = Path(base_dir)

    if not root.is_dir():
        print(f"Directory not found: {base_dir}")
        print("Set STERLING_DOCS_DIR to your knowledge directory.")
        sys.exit(1)

    db_path = str(root / "chromadb")
    collection_name = root.name

    # Find all files
    all_files = [
        f
        for f in root.rglob("*")
        if f.is_file()
        and f.suffix.lower() in ALL_EXTENSIONS
        and f.stat().st_size < MAX_FILE_SIZE
        and "chromadb" not in str(f)
        and "lancedb" not in str(f)
        and "node_modules" not in str(f)
        and ".ingest-staging" not in str(f)
    ]

    print(f"Found {len(all_files)} documentation files in {base_dir}")
    by_ext: dict[str, int] = {}
    for f in all_files:
        ext = f.suffix.lower()
        by_ext[ext] = by_ext.get(ext, 0) + 1
    for ext in sorted(by_ext):
        print(f"  {ext}: {by_ext[ext]}")

    if not all_files:
        print("No documentation files found.")
        print(f"Supported: {', '.join(sorted(ALL_EXTENSIONS))}")
        sys.exit(1)

    # Read and chunk all files
    print("\nReading and chunking files...")
    t0 = time.time()
    all_chunks: list[str] = []
    all_ids: list[str] = []
    all_meta: list[dict[str, str]] = []
    skipped = 0

    for i, f in enumerate(all_files):
        content = _read_file(f, root)
        if not content:
            skipped += 1
            continue
        chunks = _chunk_text(content)
        for j, chunk in enumerate(chunks):
            all_chunks.append(chunk)
            rel_path = str(f.relative_to(root)).replace("/", "_").replace("\\", "_")
            all_ids.append(f"{rel_path}-{j}")
            all_meta.append({"source": str(f.relative_to(root)), "file": f.name})

        if (i + 1) % 1000 == 0:
            print(f"  Read {i + 1}/{len(all_files)} files ({len(all_chunks)} chunks)")

    read_time = time.time() - t0
    print(f"  {len(all_chunks)} chunks from {len(all_files) - skipped} files ({skipped} skipped)")
    print(f"  Read time: {read_time:.1f}s")

    # Load embedding model (ONNX for CPU speed)
    print("\nLoading embedding model (ONNX)...")
    t0 = time.time()
    from sentence_transformers import SentenceTransformer  # noqa: PLC0415

    model = SentenceTransformer("all-MiniLM-L6-v2", backend="onnx")
    print(f"  Model loaded in {time.time() - t0:.1f}s")

    # Batch embed
    print(f"\nEmbedding {len(all_chunks)} chunks (batch_size={EMBED_BATCH_SIZE})...")
    t0 = time.time()
    embeddings = model.encode(
        all_chunks,
        batch_size=EMBED_BATCH_SIZE,
        show_progress_bar=True,
        normalize_embeddings=True,
    )
    embed_time = time.time() - t0
    chunks_per_sec = len(all_chunks) / embed_time if embed_time > 0 else 0
    print(f"  Embedded in {embed_time:.1f}s ({chunks_per_sec:.0f} chunks/sec)")

    # Store in ChromaDB
    print(f"\nStoring in ChromaDB at {db_path}...")
    t0 = time.time()
    import chromadb  # noqa: PLC0415

    client = chromadb.PersistentClient(path=db_path)

    if args.reset:
        try:
            client.delete_collection(collection_name)
            print(f"  Deleted existing collection '{collection_name}'")
        except Exception:
            pass

    collection = client.get_or_create_collection(
        name=collection_name,
        metadata={"hnsw:space": "cosine"},
    )

    # Batch insert (ChromaDB has a max batch size)
    max_batch = 5000
    for start in range(0, len(all_chunks), max_batch):
        end = min(start + max_batch, len(all_chunks))
        collection.add(
            ids=all_ids[start:end],
            documents=all_chunks[start:end],
            embeddings=embeddings[start:end].tolist(),
            metadatas=all_meta[start:end],
        )
        print(f"  Stored {end}/{len(all_chunks)}")

    store_time = time.time() - t0
    print(f"  Stored in {store_time:.1f}s")

    total_time = read_time + embed_time + store_time
    print(f"\nIngestion complete in {total_time:.1f}s ({total_time / 60:.1f} min)")
    print(f"  {len(all_chunks)} chunks from {len(all_files)} files")
    print(f"  ChromaDB at: {db_path}")
    print(f"  Collection: {collection_name}")


if __name__ == "__main__":
    main()
